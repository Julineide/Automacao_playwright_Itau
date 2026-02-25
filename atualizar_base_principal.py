import os
from time import sleep
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.cell.cell import Cell

def _safe_save(wb, caminho, tentativas=10, espera=1.0):
    for i in range(1, tentativas + 1):
        try:
            wb.save(caminho)
            return
        except PermissionError:
            print(f"Arquivo em uso... tentando novamente ({i}/{tentativas})")
            sleep(espera)
    raise PermissionError(f"Não foi possível salvar: {caminho}")

def atualizar_base_principal(
    df_itau: pd.DataFrame,
    caminho_base_principal: str,
    caminho_export_goal: str,
    nome_aba="Base Ativos Itau",
    preencher_B_em_existentes_vazios=True  # se True, preenche B quando estiver vazio
):
    if not os.path.exists(caminho_base_principal):
        raise FileNotFoundError(f"Base principal não encontrada:\n{caminho_base_principal}")
    
    # Configurações de estilo
    estilo_fonte = Font(size=9)
    estilo_alinhamento = Alignment(horizontal="center", vertical="center")
    
    df_goal = pd.read_excel(caminho_export_goal)
    # Assume que a placa está na primeira coluna (índice 0) do Export_goal
    df_goal["Placa_Match"] = df_goal.iloc[:, 0].astype(str).str.strip().str.upper()
    goal_map = {row["Placa_Match"]: row for _, row in df_goal.iterrows()}

    # 1) Base consolidada por placa (última ocorrência do arquivo) e normalizada
    base = df_itau.copy()
    base["Placa"] = base["Placa"].astype(str).str.strip().str.upper()
    base = base.drop_duplicates(subset=["Placa"], keep="last").reset_index(drop=True)

    # dicionários
    serie_by = dict(zip(base["Placa"], base["NumeroSerie"]))
    data_by  = dict(zip(base["Placa"], base["UltimaDataHora"]))

    # 2) Abre workbook
    wb = load_workbook(caminho_base_principal)
    ws = wb[nome_aba]
    if nome_aba not in wb.sheetnames:
        raise KeyError(f"Aba '{nome_aba}' não encontrada")

    # 3) Placas existentes (normalizadas)
    placas_ws = []
    row_idx_by_placa = {}
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):
        cell = row[0]
        placa_raw = str(cell.value).strip() if cell.value else ""
        if placa_raw:
            placa = placa_raw.upper()
            placas_ws.append(placa)
            row_idx_by_placa[placa] = cell.row

    set_ws   = set(placas_ws)
    set_itau = set(base["Placa"])

    novas    = sorted(list(set_itau - set_ws))
    inativas = sorted(list(set_ws - set_itau))

    # 4) Remove inativas (de baixo pra cima)
    if inativas:
        rows_to_del = []
        for row in ws.iter_rows(min_row=2, max_col=1):
            val = row[0].value
            placa = str(val).strip().upper() if val else ""
            if placa in inativas:
                rows_to_del.append(row[0].row)
        for r in sorted(rows_to_del, reverse=True):
            ws.delete_rows(r, 1)

    # 5) Reindexa após remoções
    placas_ws = []
    row_idx_by_placa = {}
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):
        cell = row[0]
        placa_raw = str(cell.value).strip() if cell.value else ""
        if placa_raw:
            placa = placa_raw.upper()
            placas_ws.append(placa)
            row_idx_by_placa[placa] = cell.row

    # 6) Inserir NOVAS (A, B, C)
    for placa in novas:
            num = serie_by.get(placa, None)
            dt  = data_by.get(placa, None)
            ws.append([placa, num, dt])
            last = ws.max_row
            
            # Aplicar estilo nas novas células A, B, C
            for col in range(1, 4):
                cell = ws.cell(row=last, column=col)
                cell.font = estilo_fonte
                cell.alignment = estilo_alinhamento

    # 7) Atualizar TODA a coluna C com o valor da Itaú_base
    #    (replicar, não calcular)
    row_idx_by_placa = {}
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):
        c = row[0]
        placa = str(c.value).strip().upper() if c.value else ""
        if placa:
            row_idx_by_placa[placa] = c.row
            # Atualiza Data VDO
            if placa in data_by:
                cellC: Cell = ws.cell(row=c.row, column=3)
                cellC.value = data_by[placa]
                cellC.number_format = "dd/mm/yyyy"
                cellC.font = estilo_fonte
                cellC.alignment = estilo_alinhamento

    # 8) Preencher B (Série) se vazio
    if preencher_B_em_existentes_vazios:
        for placa, row_idx in row_idx_by_placa.items():
            if placa in serie_by:
                cellB = ws.cell(row=row_idx, column=2)
                if cellB.value in (None, "", " "):
                    cellB.value = serie_by[placa]
                    cellB.font = estilo_fonte
                    cellB.alignment = estilo_alinhamento

    # 9) Preencher Colunas 4 a 8 (Goal) com busca por placa e formatação
    for placa, row_idx in row_idx_by_placa.items():
        if placa in goal_map:
            row_goal = goal_map[placa]
            
            mapeamento_goal = [
                (4, 3, None),               # Proposta #
                (5, None, "Itau Unibanco"), # Nome conta
                (6, 4, None),               # Modelo
                (7, 2, None),               # Código Escopo
                (8, 12, None)               # Dt. atendimento
            ]

            for col_idx, goal_idx, valor_fixo in mapeamento_goal:
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if cell.value in (None, "", " "):
                    cell.value = valor_fixo if valor_fixo else row_goal.iloc[goal_idx]
                    if col_idx == 8:
                        cell.number_format = "dd/mm/yyyy"

                # Aplica formatação independentemente de ter acabado de preencher ou já existir
                cell.font = estilo_fonte
                cell.alignment = estilo_alinhamento

    _safe_save(wb, caminho_base_principal)

    print("Base atualizada com sucesso!")
    print(f"Novas placas adicionadas: {len(novas)}")
    print(f"Placas removidas: {len(inativas)}")

    # 10: Exclusão dos arquivos temporários ---
    try:
        os.remove(caminho_export_goal)
        # Se o df_itau veio de um caminho de arquivo, poderíamos excluir aqui também.
        # Mas como ele vem como DataFrame para esta função, a exclusão 
        # do arquivo da VDO deve ser feita no Main.py ou aqui se você passar o caminho.
        print("Arquivo temporário do Goal removido.")
    except Exception as e:
        print(f"Aviso: Não foi possível remover os arquivos temporários: {e}")